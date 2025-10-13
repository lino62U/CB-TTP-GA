import pandas as pd
import re
import random
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

# -------------------------
# PARSEADOR DE ARCHIVO .TXT
# -------------------------
def parse_line(line):
    """Parsea una línea de texto: (codigo, nombre, lista de (día, hora_ini, hora_fin, aula))"""
    parts = [p.strip() for p in line.strip().split('|')]
    if len(parts) < 3:
        return None
    code, name, sched = parts[0], parts[1], parts[2]
    sessions = [s.strip() for s in sched.split(',') if s.strip()]

    data = []
    for s in sessions:
        # Ejemplo: LUN_08:50_09:40@LAB3
        m = re.match(r'([A-ZÁÉÍÓÚÑ]+)_(\d{2}:\d{2})_(\d{2}:\d{2})@(\w+)', s)
        if m:
            day, start, end, room = m.groups()
            data.append((day, start, end, room))
    return code, name, data


# -------------------------
# CONSTRUCCIÓN DE TABLA
# -------------------------
def generar_horario(txt_file):
    """Crea un diccionario horario[hora][día] = 'curso @aula'"""
    with open(txt_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    horario = defaultdict(lambda: defaultdict(str))
    dias_semana = set()
    horas = set()

    for line in lines:
        parsed = parse_line(line)
        if not parsed:
            continue
        code, name, sesiones = parsed
        for day, start, end, room in sesiones:
            bloque = f"{start}-{end}"
            horario[bloque][day] += f"{name} @{room}\n"
            dias_semana.add(day)
            horas.add(bloque)

    orden_dias = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]
    dias = [d for d in orden_dias if d in dias_semana]
    horas = sorted(horas)

    data = []
    for h in horas:
        fila = {"Hora": h}
        for d in dias:
            fila[d] = horario[h].get(d, "").strip()
        data.append(fila)

    df = pd.DataFrame(data)
    return df, dias


# -------------------------
# EXPORTAR A EXCEL CON COLORES
# -------------------------
def exportar_excel(df, dias, output_xlsx="horario_coloreado.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario"

    # Esquema de colores aleatorios por curso
    curso_colores = {}

    def color_random():
        return ''.join(random.choices('89ABCDEF', k=6))  # tonos claros

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999")
    )

    # Encabezados
    headers = ["Hora"] + dias
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 25

    # Filas del horario
    for i, row in df.iterrows():
        ws.append([row.get(col, "") for col in headers])

    # Aplicar colores y formato
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            value = cell.value
            if not value:
                continue
            # identificar curso
            curso = value.split("@")[0].strip()
            if curso not in curso_colores:
                curso_colores[curso] = color_random()
            cell.fill = PatternFill("solid", fgColor=curso_colores[curso])
            cell.alignment = align_center
            cell.border = border
            cell.font = Font(size=10, color="000000")

    wb.save(output_xlsx)
    print(f"✅ Horario exportado en: {output_xlsx}")


# -------------------------
# MAIN
# -------------------------
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Genera un horario Excel coloreado (hora vs día) desde un .txt")
    parser.add_argument("--input", required=True, help="Archivo .txt con los cursos y horarios")
    parser.add_argument("--output", default="horario_coloreado.xlsx", help="Archivo de salida Excel (.xlsx)")
    args = parser.parse_args()

    df, dias = generar_horario(args.input)
    exportar_excel(df, dias, args.output)
